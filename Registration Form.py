#tkinter with  excel backend(Basant Enquiry Form)
from tkinter import *
from PIL import ImageTk, Image
from functools import partial
from tkinter import ttk
import tkinter as tk
import openpyxl,xlrd
from openpyxl import *
import pathlib

root=tk.Tk()

root.geometry("1600x1600")
root.configure(background='light blue') 
root.title("Registartion Form")

image= Image.open("besant.png")
resized_image= image.resize((225,48), Image.Resampling.LANCZOS)
img = ImageTk.PhotoImage(resized_image)
Label(root,image= img).grid(row = 0, column = 3, sticky = W, pady = 10)
Label(root, justify=LEFT, padx=10).grid(row = 0, column = 3, sticky = NE, pady = 2)




image= Image.open("oracle.png")
resized_image= image.resize((225,48), Image.Resampling.LANCZOS)
img1 = ImageTk.PhotoImage(resized_image)
Label(root,image= img1).grid(row = 2, column = 2, sticky = W, pady = 10)
Label(root, justify=CENTER ,padx=10).grid(row = 2, column = 2, sticky = W, pady = 2)



image= Image.open("microsoft.png")
resized_image= image.resize((225,48), Image.Resampling.LANCZOS)
img2 = ImageTk.PhotoImage(resized_image)
Label(root,image= img2).grid(row = 2, column = 3, sticky = W, pady = 10)
Label(root, justify=CENTER, padx=10).grid(row = 2, column = 3, sticky = W, pady = 2)




image= Image.open("psi.png")
resized_image= image.resize((225,48), Image.Resampling.LANCZOS)
img3 = ImageTk.PhotoImage(resized_image)
Label(root,image= img3).grid(row = 2, column = 4, sticky = W, pady = 10)
Label(root, justify=CENTER, padx=10).grid(row = 2, column = 4, sticky = W, pady = 2)


    

image= Image.open("pearson.png")
resized_image= image.resize((255,100), Image.Resampling.LANCZOS)
img4 = ImageTk.PhotoImage(resized_image)
Label(root,image= img4).grid(row = 2, column = 5, sticky = W, pady = 10)
Label(root, justify=CENTER, padx=10).grid(row = 4, column = 5, sticky = W, pady = 2)


username_label = ttk.Label(root, text="Name:")
username_label.grid(column=0, row=3, sticky=tk.W, padx=5, pady=5)

l1=username_entry = ttk.Entry(root)
username_entry.grid(column=1, row=3, sticky=tk.E, padx=5, pady=5)

username_label = ttk.Label(root, text="Mobile Number")
username_label.grid(column=2, row=3, sticky=tk.W, padx=5, pady=5)

l2=username_entry = ttk.Entry(root)
username_entry.grid(column=2, row=3, sticky=tk.E, padx=5, pady=5)

username_label = ttk.Label(root, text=" Email-ID")
username_label.grid(column=3, row=3, sticky=tk.W, padx=5, pady=5)

l3=username_entry = ttk.Entry(root)
username_entry.grid(column=3, row=3, sticky=tk.E, padx=5, pady=5)

username_label = ttk.Label(root, text="Qualification")
username_label.grid(column=4, row=3, sticky=tk.W, padx=5, pady=5)

l4=username_entry = ttk.Entry(root)
username_entry.grid(column=4, row=3, sticky=tk.E, padx=5, pady=5)

username_label = ttk.Label(root, text="Address")
username_label.grid(column=0, row=4, sticky=tk.W, padx=5, pady=5)

l5=username_entry = ttk.Entry(root)
username_entry.grid(column=1, row=4, sticky=tk.E, padx=5, pady=5)

username_label = ttk.Label(root, text="Course")
username_label.grid(column=2, row=4, sticky=tk.W, padx=5, pady=5)

l6=username_entry = ttk.Entry(root)
username_entry.grid(column=2, row=4, sticky=tk.E, padx=5, pady=5)

username_label = ttk.Label(root, text="Alternate No")
username_label.grid(column=3, row=4, sticky=tk.W, padx=5, pady=5)

l7=username_entry = ttk.Entry(root)
username_entry.grid(column=3, row=4, sticky=tk.E, padx=5, pady=5)



Label(root, text="Your Gender:").grid(row=5, sticky=W)
var1 = IntVar()
Checkbutton(root, text="male", variable=var1).grid(row=5,column=2, sticky=W)
var2 = IntVar()
Checkbutton(root, text="female", variable=var2).grid(row=5,column=3,sticky=W)

Label(root, text="How did you get to know us?").grid(row=6, sticky=W)
var1 = IntVar()
Checkbutton(root, text="Facebook", variable=var1).grid(row=6,column=1, sticky=W)
var2 = IntVar()
Checkbutton(root, text="Instagram", variable=var2).grid(row=6,column=2,sticky=W)
var3 = IntVar()
Checkbutton(root, text="LinkedIn", variable=var3).grid(row=6,column=3, sticky=W)
var4 = IntVar()
Checkbutton(root, text="Google", variable=var4).grid(row=6,column=4,sticky=W)
var5 = IntVar()
Checkbutton(root, text="Family/Friends", variable=var5).grid(row=6,column=5,sticky=W)

Label(root, text="Are you Experienced/Freshere?").grid(row=7, sticky=W)
var6 = IntVar()
Checkbutton(root, text="Experienced", variable=var6).grid(row=7,column=1, sticky=W)
var7 = IntVar()
Checkbutton(root, text="Fresher", variable=var7).grid(row=7,column=2, sticky=W)


Label(root, text="Do you need job assistance?").grid(row=8, sticky=W)
var8 = IntVar()
Checkbutton(root, text="Yes", variable=var8).grid(row=8,column=1, sticky=W)
var9 = IntVar()
Checkbutton(root, text="No", variable=var9).grid(row=8,column=2, sticky=W)

Label(root, text="Planning to start the course").grid(row=9, sticky=W)
var10 = IntVar()
Checkbutton(root, text="Immdiately", variable=var10).grid(row=9,column=1, sticky=W)
var11 = IntVar()
Checkbutton(root, text="After a week", variable=var11).grid(row=9,column=2, sticky=W)
var12 = IntVar()
Checkbutton(root, text="Within months", variable=var12).grid(row=9,column=3, sticky=W)
var13 = IntVar()
Checkbutton(root, text="More than a month", variable=var13).grid(row=9,column=4, sticky=W)

Label(root, text="Batch Preferred").grid(row=10, sticky=W)
var14 = IntVar()
Checkbutton(root, text="Weekdays", variable=var14).grid(row=10,column=1, sticky=W)
var15 = IntVar()
Checkbutton(root, text="Weekend", variable=var15).grid(row=10,column=2, sticky=W)
var16 = IntVar()
Checkbutton(root, text="One to One", variable=var16).grid(row=10,column=3, sticky=W)
var17 = IntVar()
Checkbutton(root, text="Fast Track", variable=var17).grid(row=10,column=4, sticky=W)


Label(root, text="Contact person from Besant Technologies").grid(row=11, sticky=W)
horizontal1 =Frame(root, bg='black', height=2,width=300)
horizontal1.place(x=285, y=425)

Label(root, text="Office use only",font="bold").grid(row=12, sticky=W)

Label(root, text="Manager").grid(row=13, sticky=W)
horizontal2 =Frame(root, bg='black', height=2,width=300)
horizontal2.place(x=65, y=475)

Label(root, text="Fees").grid(row=14, sticky=W)
horizontal2 =Frame(root, bg='black', height=2,width=300)
horizontal2.place(x=65, y=495)

Label(root, text="Comments").grid(row=15, sticky=W)
horizontal2 =Frame(root, bg='black', height=2,width=300)
horizontal2.place(x=65, y=515)
def submit():
    y=l1.get()
    z=l2.get()
    x=l3.get()
    m=l4.get()
    n=l5.get()
    o=l6.get()
    p=l7.get()
    print(y)
    print(z)
    print(x)
    print(m)
    print(n)
    print(o)
    print(p)

    file=load_workbook("Registration_Details.xlsx")
    sheet=file.active
    sheet.cell(column=1,row=sheet.max_row+1,value=y)
    sheet.cell(column=2,row=sheet.max_row,value=z)
    sheet.cell(column=3,row=sheet.max_row,value=x)
    sheet.cell(column=4,row=sheet.max_row,value=m)
    sheet.cell(column=5,row=sheet.max_row,value=n)
    sheet.cell(column=6,row=sheet.max_row,value=o)
    sheet.cell(column=7,row=sheet.max_row,value=p)

    file.save("Registration_Details.xlsx")

def cancel():
    exit()
    
def clear(): 
    l1.delete(0, END) 
    l2.delete(0, END) 
    l3.delete(0, END) 
    l4.delete(0, END) 
    l5.delete(0, END) 
    l6.delete(0, END) 
    l7.delete(0, END)
    variable.delete(0,END)

    
btn1=Button(root, text = 'Submit',command = submit,bg="blue",fg="Black").grid(row=18,column=3,sticky=W)
btn2=Button(root, text = 'Cancel',command = cancel,bg="red",fg="Black").grid(row=18,column=4,sticky=W)
btn2=Button(root, text = 'Clear',command = clear,bg="yellow",fg="Black").grid(row=18,column=5,sticky=W)

file=pathlib.Path("Registration_Details.xlsx")
if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet["A1"]="Name"
    sheet["B1"]="Mobile Number"
    sheet["C1"]="Email_id"
    sheet["D1"]="Qualification"
    sheet["E1"]="Address"
    sheet["F1"]="Course"
    sheet["G1"]="Alternate Number"

    file.save("Registration_Details.xlsx")
    

root.mainloop()

